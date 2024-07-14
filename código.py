from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import openpyxl
import time
from datetime import datetime

# Configurações do navegador
options = Options()
options.headless = False
navegador = webdriver.Firefox(options=options)
navegador.get('https://#')
wait = WebDriverWait(navegador, 10)

# Login
wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="username"]'))).send_keys("USUARIO")
time.sleep(1)
wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="password"]'))).send_keys("SENHA")
time.sleep(1)
wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="kc-login"]'))).click()
time.sleep(1)

# Função para pesquisar uma proposta
def pesquisar_proposta(numero_proposta):
    global linha_atual  # Use the global variable for the current row
    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="searchMenu"]'))).click()
    time.sleep(1)
    wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/header/div/div/div[2]/form/ul/li[18]/a/b'))).click()
    time.sleep(5)
    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tipoConsultaNContrato"]'))).click()
    time.sleep(1)
    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="numeroContrato"]'))).send_keys(numero_proposta)
    time.sleep(1)
    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="btnConsultar"]'))).click()
    time.sleep(3)

    #Clicando no botão manutenção, se houver
    try:
        wait.until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[3]/div[1]/fieldset/div[6]/div/fieldset/div[1]/table/tbody/tr/td[8]/div/a[2]'))).click()
    except NoSuchElementException:
        print(f"O botão não está presente para a proposta {numero_proposta}. Pulando para a próxima proposta.")
        return
    time.sleep(2)

    #Acessando dados do contrato
    wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/form/div/fieldset/div[1]/header/h4/a'))).click()
    time.sleep(2)

    # Captura de informações
    datafinal = navegador.find_element(By.XPATH, '//*[@id="dataFinalContrato"]').get_attribute("value")
    numerocontrato = navegador.find_element(By.XPATH,'/html/body/div[3]/form/fieldset/div/div[1]/div[1]/div/label').text
    nomedocliente = navegador.find_element(By.XPATH, '/html/body/div[3]/form/fieldset/div/div[1]/div[2]/div/label').text
    cnpj = navegador.find_element(By.XPATH, '/html/body/div[3]/form/fieldset/div/div[2]/div[2]/div/label').text
    agencia = navegador.find_element(By.XPATH, '/html/body/div[3]/form/fieldset/div/div[2]/div[4]/div/label').text
    carencia = navegador.find_element(By.XPATH, '//*[@id="qtdeCarencia"]').get_attribute("value")
    amortizacao = navegador.find_element(By.XPATH, '//*[@id="quantidadePrestacoes"]').get_attribute("value")
    valorcontrato = navegador.find_element(By.XPATH, '//*[@id="valorContrato"]').get_attribute("value")

    # Salvar informações principais no Excel
    planilha_ativa.cell(row=linha_atual, column=2).value = numerocontrato
    planilha_ativa.cell(row=linha_atual, column=3).value = agencia
    planilha_ativa.cell(row=linha_atual, column=4).value = cnpj
    planilha_ativa.cell(row=linha_atual, column=5).value = nomedocliente
    planilha_ativa.cell(row=linha_atual, column=6).value = valorcontrato
    planilha_ativa.cell(row=linha_atual, column=7).value = carencia
    planilha_ativa.cell(row=linha_atual, column=8).value = amortizacao
    planilha_ativa.cell(row=linha_atual, column=9).value = datafinal

    # Captura de informações da dilatação de prazo
    time.sleep(2)
    botaodilacao = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="linkDilacaoPrazo"]')))
    navegador.execute_script("arguments[0].click();", botaodilacao)
    time.sleep(2)

    linhas_tabela_inicial = navegador.find_elements(By.XPATH,'/html/body/div[3]/form/div/fieldset/div[11]/div/div[1]/table/tbody/tr')
    if linhas_tabela_inicial:
        ultima_linha_inicial = linhas_tabela_inicial[-1]
        informacoes_ultima_linha_inicial = ultima_linha_inicial.text
        print("Informações da última linha da Tabela Inicial:")
        print(informacoes_ultima_linha_inicial)

        # Dividir e salvar informações da tabela inicial no Excel
        info_inicial = informacoes_ultima_linha_inicial.split(' ', 3)
        planilha_ativa.cell(row=linha_atual, column=10).value = info_inicial[0]  # Informação 1
        planilha_ativa.cell(row=linha_atual, column=11).value = info_inicial[1]  # Informação 2
        planilha_ativa.cell(row=linha_atual, column=12).value = info_inicial[2]  # Informação 3
        planilha_ativa.cell(row=linha_atual, column=13).value = info_inicial[3]  # Informação 4

    else:
        print("Não foram encontradas linhas na tabela inicial.")

    # Clicar em simulação de prazo
    time.sleep(2)
    botaosimulacao = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="btnSimularDilacaoPrazo"]')))
    navegador.execute_script("arguments[0].click();", botaosimulacao)

    linhas_tabela_final = navegador.find_elements(By.XPATH,'/html/body/div[3]/form/div/fieldset/div[11]/div/div[3]/table/tbody/tr')
    if linhas_tabela_final:
        ultima_linha_final = linhas_tabela_final[-1]
        informacoes_ultima_linha_final = ultima_linha_final.text
        print("Informações da última linha da Simulação de Prazo:")
        print(informacoes_ultima_linha_final)

        # Dividir e salvar informações da simulação de prazo no Excel
        info_final = informacoes_ultima_linha_final.split(' ', 3)
        planilha_ativa.cell(row=linha_atual, column=14).value = info_final[0]  # Informação 1
        planilha_ativa.cell(row=linha_atual, column=15).value = info_final[1]  # Informação 2
        planilha_ativa.cell(row=linha_atual, column=16).value = info_final[2]  # Informação 3
        planilha_ativa.cell(row=linha_atual, column=17).value = info_final[3]  # Informação 4

    else:
        print("Não foram encontradas linhas na simulação de prazo.")

    # Fechar caixa de mensagem
    time.sleep(1)
    wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/button[1]'))).click()
    time.sleep(1)

    # Salvar a data da consulta na coluna 'L'
    data_consulta = datetime.now().strftime('%d/%m/%Y')
    planilha_ativa.cell(row=linha_atual, column=18).value = data_consulta

    # Salvar arquivo Excel
    planilha.save("lista_propostas.xlsx")

    # Criar backup a cada 10 propostas lidas
    if (linha_atual - 1) % 10 == 0:
        planilha.save("backup.xlsx")
        print(f"Backup criado: backup.xlsx")

# Leitura do arquivo Excel
arquivo_excel = "lista_propostas.xlsx"
planilha = openpyxl.load_workbook(arquivo_excel)
planilha_ativa = planilha.active

# Itera sobre as propostas na coluna A
for linha_atual in range(2, planilha_ativa.max_row + 1):
    numero_proposta = planilha_ativa.cell(row=linha_atual, column=1).value
    if numero_proposta:
        data_consulta = planilha_ativa.cell(row=linha_atual, column=12).value
        if data_consulta is None:
            pesquisar_proposta(numero_proposta)

# Fechar navegador ao finalizar
navegador.quit()
